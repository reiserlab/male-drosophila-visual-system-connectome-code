""" Excel Exporter module. """
import os
from pathlib import Path
import datetime
import warnings
import pandas as pd
from openpyxl.styles import Font

from .exporter import Exporter


class ExcelExporter(Exporter):
    """
    ExcelExporter is a convenient way to export a pandas dataFrame to an Excel file (xlsx).

    The implementation is very specialized at the moment, can become more generic if necessary.

    Methods
    -------
    export(dataFrame)
        Exports the pandas dataframe to a file
    """
    __filename = None
    __basename = None

    def __init__(
        self,
        output_path=None,
        output_basename=None,
        use_timestamp=True
    ):
        """Initiate the ExcelExporter

        Parameters
        ----------
        output_path : pathlib.Path, optional
            location for the generated file (Default: `os.getcwd()`)

        output_basename : str, optional
            gives a hint about the filename and sheet description (Default: 'generic')

        use_timestamp : bool, optional
            use a timestamp for the output file. (default: True)
        """
        if not output_path:
            warnings.warn("No output path provided, will write to " + os.getcwd())
            output_path = os.getcwd()
        if not output_path.exists():
            output_path.mkdir()
        today_str = ""
        if not output_basename:
            output_basename = "generic"
        self.__basename = output_basename
        tmp_out = output_basename.lower()
        tmp_out = tmp_out.replace(" ", "-")
        if use_timestamp:
            today_str = datetime.datetime.today().strftime("%Y-%m-%dT%H-%M-%S")
            today_str += "_"
        self.__filename = Path(output_path, f"{today_str}{tmp_out}_named.xlsx")

    def export(self, data_frame):
        """
        export a data frame to a file

        Parameters
        ----------
        data_frame : pandas.DataFrame
            data frame to be stored
        """
        if not isinstance(data_frame, pd.DataFrame):
            warnings.warn("The provided dataFrame is not a valid pandas dataFrame. Expect errors.")

        with pd.ExcelWriter(self.__filename) as writer:
            data_frame.to_excel(writer, sheet_name=self.__basename)
            workbook = writer.book
            worksheet = workbook.active
            for rownum in range(2, worksheet.max_row+1):
                for colname in ['C', 'D', 'E']:
                    worksheet[f"{colname}{rownum}"].number_format = "0%"  # or "0.00%"?
            worksheet.append({
                'A': "GRAND TOTAL:",
                'B': f"=SUM(B1:B{worksheet.max_row})",
                'F': f"=SUM(F1:F{worksheet.max_row})",
                'G': f"=SUM(G1:G{worksheet.max_row})",
                'H': f"=SUM(H1:H{worksheet.max_row})",
                'I': f"=SUM(I1:I{worksheet.max_row})"})
            gt_row = worksheet.row_dimensions[worksheet.max_row]
            gt_row.font = Font(italic=True, color="FF6600")
            worksheet[f"A{worksheet.max_row}"].font = Font(
                bold=True, italic=True, color="FF6600")
        if Path(self.__filename).exists():
            print(f"successfully exported {self.__filename}")
